import openpyxl as xl
import matplotlib as plt
import os


def clear_terminal():

    # clear terminal and print nasa mission always
    os.system('cls' if os.name == 'nt' else 'clear')
    print("**********NASA-MISSIONS**************")


clear_terminal()
print("\nWelcome to NASA-Missions! (An unnofficial project)\n")
decision = ""
data_change_decision = ""

# wb = xl.load_workbook('NASA-Missions.xlsx')
while not decision == 'exit':

    clear_terminal()
    print("Would you like to create a new mission or do you want to load a mission? \n")
    decision = input(
        "Type '0' for new mission, \nType '1' for loading a mission, \nType 'help' to get information about this, \nType 'exit' to leave \n")

    # Create new mission
    if decision == '0':
        clear_terminal()

        is_planet_moon = input("\nIs it a planet or a moon? (p/m) ")

        if is_planet_moon.lower() == "p":

            clear_terminal()
            print('Creating Planet Mission \n')

            planet = {}
            planet['Planet Name'] = input(
                "What is the name of the planet you found? ")
            planet['Mission Name'] = input(
                "What is the name of the mission? (Be creative) ")
            planet['Gravity'] = input(
                "How fast do objects fall? (Rate of Gravity) ")
            planet['Extra Info'] = input("Anything else you want to add? ")

            clear_terminal()

            print("Here's the data you created: ")
            print(
                f"Planet Name: {planet['Planet Name']}\nMission Name: {planet['Mission Name']}\nGravity: {planet['Gravity']}\nExtra Info: {planet['Extra Info']}\n")
            is_correct = input("Is this correct? (y/n) ")
            print(is_correct)

            if is_correct == "n":

                while not data_change_decision == '4':
                    clear_terminal()

                    print("What would you like to change? ")
                    data_change_decision = input(
                        "0 for Planet Name\n1 for Mission Name\n2 for Gravity\n3 for Extra Info\n4 means data is correct\n")

                    if data_change_decision == '0':
                        planet['Planet Name'] = input(
                            "Type the name of the planet: ")

                    elif data_change_decision == '1':
                        planet["Mission Name"] = input(
                            "Type the name of the mission name: ")

                    elif data_change_decision == '2':
                        planet['Gravity'] = input("Type the rate of gravity: ")

                    elif data_change_decision == '3':
                        planet['Extra Info'] = input("Type the extra info: ")

                    elif data_change_decision == '4':
                        clear_terminal()
                        print("Here's the data you created: ")
                        print(
                            f"Planet Name: {planet['Planet Name']}\nMission Name: {planet['Mission Name']}\nGravity: {planet['Gravity']}\nExtra Info: {planet['Extra Info']}\n")
                        is_correct = input("Is this correct? (y/n) ")

                        # reset data decision, otherwise while loop will stop
                        data_change_decision = ""

                        if is_correct == 'y':
                            break
                        else:
                            pass

                    else:
                        input("Invalid answer, try again! (press enter to continue) ")

            # sheet = wb['Created Missions']

            # sheet.cell(2, 1).value = planet['Mission Name']
            # sheet.cell(2, 2).value = planet['Planet Name']
            # sheet.cell(2, 3).value = planet['Gravity']
            # sheet.cell(2, 4).value = planet['']

        elif is_planet_moon.lower() == "m":

            clear_terminal()
            print("Creating Moon Mission \n")

            moon = {}
            moon['Moon Name'] = input(
                "What is the name of the moon you found? (Be creative) ")
            moon['Mission Name'] = input("What is the name of the mission? ")
            moon['Gravity'] = input(
                "How fast do objects fall? (Rate of Gravity) ")
            moon['Extra Info'] = input("Anything else you want to add? ")

            clear_terminal()
            print("Here's the data you created: ")
            print(
                f"Moon Name: {moon['Moon Name']}\nMission Name: {moon['Mission Name']}\nGravity: {moon['Gravity']}\nExtra Info: {moon['Extra Info']}\n")
            is_correct = input("Is this correct? (y/n) ")

            if is_correct == "n":

                while not data_change_decision == '4':
                    clear_terminal()
                    print("What would you like to change? ")
                    data_change_decision = input(
                        "0 for Moon Name\n1 for Mission Name\n2 for Gravity\n3 for Extra Info\n4 means data is correct\n")

                    if data_change_decision == '0':
                        moon['Moon Name'] = input(
                            "Type the name of the planet: ")

                    elif data_change_decision == '1':
                        moon["Mission Name"] = input(
                            "Type the name of the mission name: ")

                    elif data_change_decision == '2':
                        moon['Gravity'] = input("Type the rate of gravity: ")

                    elif data_change_decision == '3':
                        moon['Extra Info'] = input("Type the extra info: ")

                    elif data_change_decision == '4':
                        clear_terminal()
                        print("Here's the data you created: ")
                        print(
                            f"Moon Name: {moon['Moon Name']}\nMission Name: {moon['Mission Name']}\nGravity: {moon['Gravity']}\nExtra Info: {moon['Extra Info']}\n")
                        is_correct = input("Is this correct? (y/n) ")

                        # reset data_change_decision, otherwise while loop will stop
                        data_change_decision = ''

                        if is_correct == 'y':
                            break
                        else:
                            pass

                    else:
                        input("Invalid answer, try again! (press enter to continue) ")

        else:
            clear_terminal()

            # input so that the user actually reads the error
            input('Invalid answer, try again! (enter to continue) ')

    elif not decision == 'exit':
        clear_terminal()

        # input so that the user actually reads the error
        input('Invalid decision, try again! (enter to continue) ')

else:
    print("Program ended.")
    print("*************************\n")

#  Planet should store:
# - gravity
# - distance from earth
# - number of moons
# - fuel
# - graph fuel
# = compare
# create missions to existing planets/moons in solar system
